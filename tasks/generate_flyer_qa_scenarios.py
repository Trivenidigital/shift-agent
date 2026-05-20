"""Generate Flyer Studio QA scenario workbook.

Outputs `tasks/flyer-studio-qa-scenarios.xlsx` with a Master sheet and per-area
sheets. Scenarios derived from PRs #100, #101, #102 and the deployed code on
branch `codex/flyer-edit-pipeline`.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "ID", "Area", "Scenario", "Preconditions", "Steps",
    "Expected Result", "Priority", "Type", "Channel", "Notes",
]

AREAS = [
    ("A1", "Onboarding"),
    ("A2", "Text Mode + Starter Briefs"),
    ("A3", "Image / Reference Scope"),
    ("A4", "Active Project / Revisions / Source-Edit"),
    ("A5", "Guest Orders"),
    ("A6", "Admin Dashboard"),
    ("A7", "cf-router Routing"),
]


def row(idx, area_code, area_name, scenario, pre, steps, expected, prio, typ, chan, notes):
    return [
        f"FS-{area_code}-{idx:03d}", area_name, scenario, pre, steps, expected,
        prio, typ, chan, notes,
    ]


def onboarding_scenarios():
    A, name = "A1", "Onboarding"
    out = []
    i = 1
    out.append(row(i, A, name,
        "Owner sends 'set up Flyer Studio' to start fresh onboarding",
        "Empty FlyerCustomerStore; sender phone X1 not in any account; WhatsApp inbound enabled",
        "1. Send 'set up Flyer Studio' from phone X1.",
        "Receives Welcome message: 'Flyer Studio\\n------------\\nWelcome. I can set up your flyer account here on WhatsApp.\\n\\n{plan_lines}\\n\\nFirst, what is your business name?'. Session created at status=collecting_business_name.",
        "P0", "Happy", "WhatsApp",
        "Verify state file under state/flyer/customers.json includes new onboarding_session keyed to chat_id."
    )); i += 1
    out.append(row(i, A, name,
        "Owner taps trial CTA button 'Help me create a beautiful flyer for my business'",
        "Empty store; campaign CTA wired for trial",
        "1. Customer taps WhatsApp button with CTA text.\n2. cf-router intercepts via flyer_campaign_cta_text.",
        "Welcome reply prefixed with: 'Absolutely, lets create a beautiful flyer for your business. I will set up your free trial first... Your free trial includes 3 free sample flyers.' followed by name prompt. session.plan_id is set to 'trial' and choosing_plan step is skipped.",
        "P0", "Happy", "WhatsApp",
        "Trial trigger keywords: free trial, start trial, try free, 3 free, set up flyer studio, act now! save time and money."
    )); i += 1
    out.append(row(i, A, name,
        "Owner taps 'Create One Flyer - $4' Quick Flyer CTA",
        "Empty store; campaign CTA wired",
        "1. Tap the $4 quick flyer button (delivers a CTA containing '$4', 'one flyer', or 'quick flyer').",
        "Routed via _try_flyer_campaign_cta_intercept; no onboarding wizard triggered for paid-per-flyer path; pricing intake initiated under guest-order flow.",
        "P0", "Happy", "WhatsApp",
        "is_quick_flyer_campaign_cta should classify TRUE; downstream uses manage-flyer-guest-order start path."
    )); i += 1
    out.append(row(i, A, name,
        "Owner taps subscription CTA",
        "Empty store",
        "1. Tap subscription CTA button.",
        "Standard onboarding starts; plan_id stays unset until choosing_plan step (choices include starter/growth/unlimited).",
        "P1", "Happy", "WhatsApp",
        "Verify choosing_plan prompt enumerates configured subscription plans."
    )); i += 1
    out.append(row(i, A, name,
        "Happy path: complete trial onboarding end-to-end",
        "Trial CTA tapped; no prior customer record",
        "1. Reply business name 'Triveni Cafe'.\n2. Address: '123 Main St, Dallas TX'.\n3. Public phone: '+1 555 010 2000'.\n4. Business WhatsApp: '+1 555 010 2000' or SKIP.\n5. Authorized request number: same or SKIP.\n6. Business profile: 'Indian restaurant, English and Telugu'.\n7. Reply CONFIRM at summary.",
        "Each prompt advances the state machine. Final reply: 'Free trial active for {customer_id}. You have 3 free sample flyers. Send your first flyer request now.' Customer record persisted with status='trial' and plan_id='trial'.",
        "P0", "Happy", "WhatsApp",
        "Confirm flyer_customer.status transitions pending -> trial."
    )); i += 1
    out.append(row(i, A, name,
        "Happy path: paid plan onboarding ends in payment_pending",
        "Subscription CTA; no prior record",
        "1. Complete all profile fields.\n2. Choose paid plan (e.g. 'starter').\n3. Send CONFIRM.",
        "Reply: 'Registration saved as {customer_id} on the starter plan. Pay here: {url}'. Customer status=payment_pending; checkout URL populated.",
        "P0", "Happy", "WhatsApp",
        "If url missing, expect fallback: 'Payment link is pending. We will send a secure Stripe/Razorpay link shortly.'"
    )); i += 1
    out.append(row(i, A, name,
        "Validation: empty business name",
        "Onboarding session at collecting_business_name",
        "1. Reply with empty string or whitespace.",
        "Reply: 'Flyer Studio\\n------------\\nPlease send the business name.\\n\\nFirst, what is your business name?'. State unchanged.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Validation: address missing digit and signal words",
        "Onboarding at collecting_business_address",
        "1. Reply 'My Place'.",
        "Reply prefixed with: 'Please send the full business address, including street/city/state if available.' Status unchanged.",
        "P1", "Negative", "WhatsApp",
        "Acceptable signals: street/road/ave/blvd/state abbreviations (nc|sc|fl|tx|va|md|oh|ca|ny|nj) or any digit."
    )); i += 1
    out.append(row(i, A, name,
        "Validation: invalid phone format",
        "Onboarding at collecting_public_phone",
        "1. Reply 'abc 12'.",
        "Reply prefixed with: 'Please send a valid phone number with country code, or a US 10-digit number.'",
        "P1", "Negative", "WhatsApp",
        "Both E.164 and bare 10-digit US are accepted."
    )); i += 1
    out.append(row(i, A, name,
        "SKIP at collecting_business_whatsapp after public phone saved advances state",
        "Onboarding at collecting_business_whatsapp; public_phone already saved",
        "1. Reply 'SKIP' at collecting_business_whatsapp.",
        "business_whatsapp_number stays unset (None); status -> collecting_authorized_request_number; next prompt sent ('What is the authorized flyer request number?').",
        "P1", "Happy", "WhatsApp",
        "Round-2 rewrite (2026-05-19): original precondition 'before public phone collected' was structurally unreachable in the state machine — collecting_business_whatsapp only follows collecting_public_phone."
    )); i += 1
    out.append(row(i, A, name,
        "Validation: language-only business profile reply",
        "Onboarding at collecting_business_profile",
        "1. Reply 'English'.",
        "ValueError raised. Reply prefixed with: 'Please include the business type, for example: Hair salon, English.'",
        "P1", "Negative", "WhatsApp",
        "Verifies _parse_profile_text no-fallback behavior added in PR 102."
    )); i += 1
    out.append(row(i, A, name,
        "EDIT NAME from summary",
        "Onboarding at confirming_summary",
        "1. Reply 'EDIT NAME: Triveni Coffee'.",
        "Business name updated to 'Triveni Coffee'; summary re-rendered showing new value; status stays confirming_summary.",
        "P0", "Happy", "WhatsApp",
        "Repeat for ADDRESS, PHONE, WHATSAPP, AUTHORIZED, PROFILE, PLAN."
    )); i += 1
    out.append(row(i, A, name,
        "Unknown EDIT field",
        "Confirming summary",
        "1. Reply 'EDIT FOO: bar'.",
        "Reply: 'Unknown edit field. Use EDIT NAME, ADDRESS, PHONE, WHATSAPP, AUTHORIZED, PROFILE, or PLAN.'",
        "P2", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Non-CONFIRM reply at summary",
        "Confirming summary",
        "1. Reply 'ok looks good'.",
        "Reply: 'Reply CONFIRM to finish registration, or send EDIT FIELD: value.' Status unchanged.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "BACK navigation steps backward (paid + trial paths differ at summary)",
        "Onboarding at collecting_public_phone (name+address saved) OR at confirming_summary (trial vs paid path)",
        "1. From collecting_public_phone reply 'BACK'.\n2. Separately, from confirming_summary reply 'BACK' once on a trial session and once on a paid session.",
        "From collecting_public_phone: status returns to collecting_business_address; prior address cleared; address prompt re-sent. From confirming_summary (trial path, plan_id='trial'): status returns to collecting_business_profile; plan_id stays 'trial'; choosing_plan is correctly skipped. From confirming_summary (paid path): status returns to choosing_plan; plan_id cleared.",
        "P1", "Edge", "WhatsApp",
        "Round-2 update (2026-05-19): pinned by tests/test_flyer_onboarding.py::test_trial_back_from_confirming_summary_skips_choosing_plan + test_paid_back_from_confirming_summary_returns_to_choosing_plan after BUG-FLYER-QA-2026-05-19-001 fix."
    )); i += 1
    out.append(row(i, A, name,
        "RESTART resets session, preserves pending brand assets",
        "Onboarding partially complete; logo image previously uploaded and saved as pending asset",
        "1. Reply 'RESTART'.",
        "Status returns to collecting_business_name; all field values cleared; pending_brand_assets retained.",
        "P1", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "HELP returns current prompt unchanged",
        "Any onboarding step",
        "1. Reply 'HELP'.",
        "Current step prompt re-sent; state unchanged.",
        "P2", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Duplicate sender: phones match + name fuzzy-match >=0.86",
        "Existing active customer 'Triveni Supermarket' with sender X1 authorized",
        "1. From phone X1, start onboarding fresh.\n2. Reply with business name 'Triveni Super Market'.",
        "Session discarded; sender phone added to existing account's authorized_request_numbers; reply: 'This number is already set up for Triveni Supermarket.\\n\\nYou can start creating a flyer now.'",
        "P0", "Edge", "WhatsApp",
        "Tests _find_named_duplicate_customer + _connect_recovered_sender."
    )); i += 1
    out.append(row(i, A, name,
        "Duplicate sender: phone match, different unrelated name",
        "Existing customer 'Acme Realty' owns phone X1",
        "1. From phone X1, start onboarding fresh and provide business name 'Joes Diner'.",
        "Reply: 'That phone number belongs to another Flyer Studio account.\\n\\nReply EDIT WHATSAPP or EDIT AUTHORIZED with a different number.' Status stays at confirming_summary.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Brand asset uploaded mid-onboarding (logo)",
        "Onboarding at collecting_business_profile",
        "1. Upload an image (PNG) with caption 'our logo'.",
        "Image classified as 'logo' (caption matches). Stored under pending_brand_assets keyed by kind='logo'. Reply: 'Logo saved. I will attach it to this account during onboarding.\\n\\n{next_prompt}'.",
        "P0", "Happy", "WhatsApp",
        "Confirm copy stored at /opt/shift-agent/state/flyer/brand_assets/<owner_key>/..."
    )); i += 1
    out.append(row(i, A, name,
        "Brand asset uploaded mid-onboarding (template/sample)",
        "Onboarding at any step",
        "1. Upload image with caption 'sample flyer'.",
        "Classified as 'template' (caption matches sample/reference/menu/poster regex). Reply: 'Template saved...' followed by next prompt.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Non-admin tries to replace logo on active account",
        "Active customer A; sender X2 not in authorized_request_numbers and not business_whatsapp_number",
        "1. From X2, upload a new logo image.",
        "Reply: 'Only the business WhatsApp number or account owner can replace saved logos/templates for this account.' No replacement performed.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Admin successfully replaces logo on active account",
        "Active customer A; sender is business_whatsapp_number",
        "1. Upload new logo image.",
        "Reply: 'Logo saved and will be used for future flyers.' Previous logo replaced.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "LID-only sender (no resolved phone) completes onboarding",
        "Sender chat resolved via identify-sender to LID with no phone mapping",
        "1. Complete all onboarding steps via LID-only inbound.",
        "Session keyed by chat_id with sender_phone=None; saves successfully; downstream find_session lookup uses chat_id fallback.",
        "P1", "Edge", "WhatsApp",
        "Confirm safe_io atomic_write_text used for state."
    )); i += 1
    out.append(row(i, A, name,
        "Wrong-field reply during onboarding stays in repair loop (not LLM)",
        "Onboarding at collecting_business_address",
        "1. Reply with business name again 'Triveni Cafe'.",
        "Stays in onboarding repair; does NOT escape to generic LLM. Validation error prompt re-sent.",
        "P0", "Edge", "WhatsApp",
        "Lesson logged in PR 101; verify no Hermes fallthrough."
    )); i += 1
    out.append(row(i, A, name,
        "Compound CONFIRM with trailing flyer request: 'CONFIRM. Create flyer for diwali special'",
        "Onboarding at confirming_summary; trial plan",
        "1. Reply 'CONFIRM. Create flyer for diwali special.'",
        "Onboarding completes -> 'Free trial active...'. Starter brief suppression triggers (_suppress_flyer_starter_brief). Trailing 'Create flyer for diwali special' is routed to primary intercept; project intake begins.",
        "P0", "Edge", "WhatsApp",
        "Confirms PR 101 will_route_trailing logic + PR 102 starter-brief suppression."
    )); i += 1
    return out


def text_mode_scenarios():
    A, name = "A2", "Text Mode + Starter Briefs"
    out, i = [], 1
    out.append(row(i, A, name,
        "Trial-active customer 'Create flyer' lands on starter brief intercept",
        "Active trial customer; business_category='restaurant'",
        "1. Send 'Create flyer'.",
        "is_vague_flyer_start=True. cf-router sends restaurant starter brief: 'Flyer Studio\\n------------\\nHere is a starter flyer request.\\nBusiness: {name}\\nEdit anything below and send it back.\\n\\nCreate a professional flyer for my restaurant.\\n\\nMain heading:\\nWeekend Specials...' Returns action=skip.",
        "P0", "Happy", "WhatsApp",
        "Reason='flyer_starter_brief' written to audit log."
    )); i += 1
    out.append(row(i, A, name,
        "Active customer 'Make flyer' (vague) routes via starter brief",
        "Active paid customer; category='salon_beauty'",
        "1. Send 'Make flyer'.",
        "Salon/beauty starter brief sent: 'Create a stylish flyer for my salon or beauty business... Fresh Look, Beautiful Confidence...'",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Active customer 'Create flyer for my restaurant special'",
        "Active customer",
        "1. Send 'Create flyer for my restaurant special'.",
        "has_detail=True (matches 'special'); is_vague_flyer_start=False; routed normally to primary intercept; project created.",
        "P0", "Happy", "WhatsApp",
        "Detail tokens: $, :, weekday/today/tomorrow/weekend, sale/offer/discount/special/menu/event/seo/grand opening/class."
    )); i += 1
    out.append(row(i, A, name,
        "Each of 10 categories matches its starter brief",
        "Active customers each in one of: restaurant, grocery, digital_marketing_agency, salon_beauty, realtor, tutor_education, event_planner, tax_accounting, temple_nonprofit, home_services",
        "1. From each, send 'Create flyer'.",
        "Each receives the category-specific brief verbatim (verify main heading + first paragraph).",
        "P0", "Happy", "WhatsApp",
        "Per starter_briefs.py; word-boundary keyword match used (regex)."
    )); i += 1
    out.append(row(i, A, name,
        "Unknown business category falls back to local_business brief",
        "Active customer with business_category='widget repair'",
        "1. Send 'Create flyer'.",
        "Receives local_business brief: heading 'Special Offer'.",
        "P1", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "AI variant: marketing agency mentioning AI gets ai_body",
        "Active customer with business_category='AI marketing agency'",
        "1. Send 'Create flyer'.",
        "ai_body of digital_marketing_agency brief used: heading 'Grow Your Business with AI-Powered Marketing'.",
        "P1", "Happy", "WhatsApp",
        "Triggered by regex on category text: '(^| )ai(?: |$)|artificial intelligence|ai marketing|ai-powered'."
    )); i += 1
    out.append(row(i, A, name,
        "Customer edits brief and replies",
        "After starter brief sent",
        "1. Reply with edited brief text including specific items, prices, dates.",
        "Treated as flyer request -> primary intercept creates project; status transitions to collecting_required_info or awaiting_assets.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Text Mode ready (post-onboarding) appends starter brief",
        "Trial customer just activated; chooses Text Mode in intake",
        "1. Intake reaches mode-choice; reply '2'.",
        "'Text Mode is ready in English. Send your flyer request... {starter_brief_message}'. Brief uses business_category from customer record.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Guided Mode prompts: goal -> schedule -> items -> location -> assets",
        "Active trial customer in intake; selects guided mode",
        "1. At choosing_mode reply '1'.\n2. Reply each prompt with valid content.\n3. Final asset reply.",
        "Sequence: 'First, what are you promoting?' -> 'What date, time, or schedule should appear...' -> 'What items, offers, prices...' -> 'What location and contact number...' (with optional 'Saved location/contact: ... USE SAVED') -> 'Any style preference...'. Final raw_request synthesized as 'Create a professional flyer. Promotion: ... Schedule: ... Items/offers: ... Location/contact: ... Style/assets: ...'",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Guided Mode 'USE SAVED' shortcut for location",
        "Guided collecting_location step; saved address + phone present on customer",
        "1. Reply 'USE SAVED'.",
        "Location step uses saved address/phone; intake advances to assets prompt.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Guided Mode 'SKIP' on optional schedule",
        "Guided collecting_schedule",
        "1. Reply 'SKIP'.",
        "Schedule field marked empty; next prompt sent (items).",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Language choice: '5' selects Tamil and updates customer profile",
        "Active customer in intake at choosing_language",
        "1. Reply '5'.",
        "Customer.preferred_language is set to 'ta' (Tamil); reply: 'Great. I will use Tamil. How would you like to create your flyer?'",
        "P1", "Happy", "WhatsApp",
        "Languages (deployed order): 1.English 2.Telugu 3.Hindi 4.Malayalam 5.Tamil 6.Kannada 7.Gujarati 8.Marathi 9.Punjabi 10.Spanish 11.Mixed/Other. Pinned by tests/test_flyer_onboarding.py::test_language_menu_pins_deployed_order_at_positions_4_through_6 (round-2 fix 2026-05-19)."
    )); i += 1
    out.append(row(i, A, name,
        "Language choice: invalid input",
        "Choosing_language",
        "1. Reply 'pirate'.",
        "Reply prefixed: 'Please choose one of these languages.' Same language menu re-sent.",
        "P2", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Mode-choice: invalid input",
        "Choosing_mode",
        "1. Reply 'maybe'.",
        "Reply: 'Please choose a creation mode.' followed by mode menu.",
        "P2", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Unregistered sender with source=start_trial routed to onboarding mid-intake",
        "Intake started for sender with no customer record",
        "1. At choosing_mode reply '1' or '2'.",
        "Intake session discarded; onboarding starts. Reply: 'I will set up your free trial first... First, what is your business name?'",
        "P1", "Edge", "WhatsApp",
        ""
    )); i += 1
    return out


def image_reference_scenarios():
    A, name = "A3", "Image / Reference Scope"
    out, i = [], 1
    out.append(row(i, A, name,
        "Active customer sends own-branded reference image",
        "Active customer with logo saved; sender authorized",
        "1. Send image of a previous flyer carrying same business name/logo.",
        "Reference-scope check returns 'authorized'; project flow proceeds with image as reference. No clarify prompt.",
        "P0", "Happy", "WhatsApp",
        "check-flyer-reference-scope script invoked via Hermes vision."
    )); i += 1
    out.append(row(i, A, name,
        "Customer sends third-party-branded reference image -> clarify",
        "Active customer",
        "1. Send a competitor's flyer image.",
        "Receives clarify prompt: 'I could not confirm whether the attached flyer belongs to {business_name}... reply with how it is connected... or use as reference...'. State saved in reference_scope_pending.json with TTL 1800s.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reference-scope option 1 (authorized)",
        "Reference-scope pending; awaiting_choice",
        "1. Reply 'option 1'.",
        "Choice consumed; status switched to awaiting_authorization_details. Reply asks customer to send authorization note.",
        "P0", "Happy", "WhatsApp",
        "_reference_scope_choice patterns: '1', 'option 1', 'authorized', 'i own', 'we own'."
    )); i += 1
    out.append(row(i, A, name,
        "Reference-scope option 2 (use as reference)",
        "Reference-scope pending; awaiting_choice",
        "1. Reply 'use as reference'.",
        "Choice consumed; project intake started using image as inspiration with raw_request prefix indicating 'reference only'. Pending entry removed.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Authorization details follow-up after option 1",
        "Reference-scope status=awaiting_authorization_details",
        "1. Reply 'We are reselling these items under license from XYZ.'",
        "authorization_note recorded on pending state; entry not removed. Project intake proceeds with authorization metadata attached.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Continue with saved business identity after option 1",
        "Reference-scope status=awaiting_authorization_details",
        "1. Reply 'use saved details'.",
        "choice='use_account_details'; authorization metadata uses customer record only.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reference scope TTL expiry (30 min)",
        "Reference-scope entry created >30min ago",
        "1. After 31 minutes reply 'option 1'.",
        "Entry pruned on read; option reply not matched; downstream classifies as generic text. No leak of expired state.",
        "P1", "Edge", "WhatsApp",
        "_read_reference_scope_state filters by expires_at."
    )); i += 1
    out.append(row(i, A, name,
        "TOCTOU safety: two concurrent option-1 replies same chat",
        "Reference-scope pending; lock acquired via _reference_scope_state_lock",
        "1. Send two simultaneous inbound messages 'option 1' from same chat.",
        "Only one consumes the pending entry; the second receives None (no double-consume); state file is atomically written.",
        "P0", "Edge", "WhatsApp",
        "Lock added in commit f41d54b."
    )); i += 1
    out.append(row(i, A, name,
        "Block decision short-circuits flow",
        "Vision returns scope=block",
        "1. Customer sends image flagged as block.",
        "Reply: 'I could not confirm this attached flyer belongs to this business account. Please send a related flyer/reference or use Create One Flyer - $4 for unrelated work.' No project created.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Exact reference edit request (own flyer + edit instruction)",
        "Authorized scope; customer sends own flyer + caption 'change price to $5'",
        "1. Send image + caption.",
        "is_exact_reference_edit_request=True. Project created with manual_edit_required=True; raw_request synthesized 'Edit uploaded flyer/source artwork. Customer requested: change price to $5'. Source-edit pipeline triggers render_source_edit_preview.",
        "P0", "Happy", "WhatsApp",
        "Source-edit verification_mode='source_edit_integrity_only'."
    )); i += 1
    out.append(row(i, A, name,
        "Source-edit preview delivery success",
        "manual_edit_required project; OPENROUTER_API_KEY configured",
        "1. Run generate-flyer-concepts on the project.",
        "Preview image sent via WhatsApp; manifest carries warning: 'Source-preserving edit output is model-edited artwork; inspect the preview visually before approval.'",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Source-edit preview generation failure -> manual ack",
        "manual_edit_required; OpenRouter source-edit failure",
        "1. Trigger generation; force failure (no key or network error).",
        "send_flyer_manual_edit_ack invoked. Customer receives manual edit acknowledgement. Project queued.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Brand-asset intercept upload (no flyer intent)",
        "Active customer, sender authorized, sends image without flyer-creation text",
        "1. Send image of new logo with caption 'logo update'.",
        "_try_flyer_brand_asset_intercept stores image under brand_assets; reply: 'Logo saved and will be used for future flyers.'",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Brand asset stored under correct owner_key path",
        "Run store-flyer-brand-asset CLI directly",
        "1. Call store-flyer-brand-asset --customer-id <id> --kind logo --media-path tmp/x.png.",
        "File copied to /opt/shift-agent/state/flyer/brand_assets/<owner_key>/<assetId>-logo.png. Decisions.log entry recorded.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reference-scope pending file uses safe_io atomic write + FileLock",
        "Trigger any save_flyer_reference_scope_pending path",
        "1. Send a third-party reference image.",
        "Inspect state/flyer/reference_scope_pending.json: written via _reference_scope_atomic_writer (safe_io.atomic_write_text). File-lock via _reference_scope_state_lock (safe_io.FileLock) held across read-modify-write.",
        "P0", "Edge", "WhatsApp",
        "Regression test for PR 101 review issue."
    )); i += 1
    return out


def active_project_scenarios():
    A, name = "A4", "Active Project / Revisions / Source-Edit"
    out, i = [], 1
    out.append(row(i, A, name,
        "Customer with active project sends revision intent 'change date to May 15'",
        "Active customer; active project in revising_design or awaiting_final_approval",
        "1. Send 'change date to May 15'.",
        "_try_flyer_active_project_intercept fires; revision patch extracted; project field date updated; new preview generated; status -> generating_concepts -> awaiting_final_approval.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Visual-only revision ('make it bigger')",
        "Active project at awaiting_final_approval",
        "1. Send 'make it bigger and more colorful'.",
        "RevisionPatchResult.visual_only=True; no field updates; regeneration with adjusted style notes.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Ambiguous price revision ('from $4 to $5') where $4 not found in project",
        "Active project without $4 in notes",
        "1. Send 'change price from $4 to $5'.",
        "ambiguous=True; unresolved_reason set; clarification reply requesting which item to update.",
        "P1", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Item swap revision",
        "Active project",
        "1. Send 'swap dosa with biryani'.",
        "Instruction appended to notes and raw_request; regeneration produced with item swap.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Headline revision",
        "Active project",
        "1. Send 'title should be Diwali Specials, not Weekend Specials'.",
        "Title field updated; regenerated preview shows new heading.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Vague new-flyer phrasing routes to starter brief, not force_new",
        "Active trial/paid customer with an active project",
        "1. Send 'start a new flyer for next week'.",
        "is_vague_flyer_start=True (no $/time/weekday/sale signals). Starter-brief intercept fires BEFORE should_start_new_flyer_over_active; reason='flyer_starter_brief'; action=skip. No force_new path is taken for vague phrasing. To force a new project, the customer must include detail signals ($, time, weekday, sale/offer/special/etc.).",
        "P0", "Happy", "WhatsApp",
        "Round-2 rewrite (2026-05-19): PR #102 introduced the starter-brief gate which intentionally short-circuits before the force_new check on vague text. Pinned by tests/test_cf_router_flyer_routing.py vague-start tests."
    )); i += 1
    out.append(row(i, A, name,
        "Delivered project receives an edit request",
        "Project status=delivered",
        "1. Send 'change time to 6 PM'.",
        "find_active_flyer_project_by_sender now treats delivered as non-terminal (PR 101); is_flyer_revision_intent matches; project enters edit/source-edit flow (manual_edit_required if no source artwork).",
        "P0", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Delivered project: vague follow-up does NOT reopen",
        "Project status=delivered",
        "1. Send 'thanks!'",
        "is_flyer_revision_intent=False; project not reopened. Message routed through generic LLM (or no-op).",
        "P1", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Missing required fields prompt",
        "Project at collecting_required_info; missing date and venue",
        "1. Run flyer_dispatcher.",
        "Reply: 'Please send the date, venue or location. I will keep the flyer copy in {Language}.'",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Concept previews delivered for approval",
        "Project at generating_concepts",
        "1. Run generate-flyer-concepts.",
        "Concept image(s) sent; status -> awaiting_final_approval. Customer can reply 'approve concept 1', 'change ...', etc.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Approval triggers finalize + delivery",
        "Project at awaiting_final_approval",
        "1. Reply 'approve concept 2'.",
        "Status -> finalizing_assets -> delivered. send-flyer-package sends PNG + PDF + markdown. flyer-delivery-report appends decisions.log entry.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Instruction-leak quality block",
        "Project has fact text 'Create a flyer for our diwali special' in heading",
        "1. Run smoke-flyer-quality.",
        "Reports blocker: 'instruction text leaked into flyer copy: heading=...'. Project does not advance to delivery.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Regional-language warning fires (no logo/template uploaded)",
        "Project with preferred_language=te (Telugu); no brand assets",
        "1. Run smoke-flyer-quality.",
        "Warning: 'regional_language_font_render_check_required'. Project advances but flagged.",
        "P2", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Customer with active project + paid guest order",
        "Active project; paid guest order open",
        "1. Send 'create another flyer'.",
        "find_paid_flyer_guest_order returns the paid order; routed with force_new=True; new project consumes the order on completion.",
        "P0", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Quota exceeded on trial -> upgrade reply",
        "Trial customer used 3/3 sample flyers",
        "1. Request another flyer.",
        "Reply: 'Your free trial has used 3/3 sample flyers. Upgrade now to keep creating professional flyers: reply CHANGE PLAN STARTER, CHANGE PLAN GROWTH, or CHANGE PLAN UNLIMITED.'",
        "P0", "Negative", "WhatsApp",
        ""
    )); i += 1
    return out


def guest_order_scenarios():
    A, name = "A5", "Guest Orders"
    out, i = [], 1
    out.append(row(i, A, name,
        "Start guest order: pending_payment reply with checkout URL",
        "No prior guest order for sender X1; price configured at 400 cents",
        "1. Trigger start_guest_order via quick_flyer CTA.",
        "Reply: 'Flyer Studio\\n------------\\nCreate one professional flyer for $4.\\nNo monthly plan. No setup required.\\n\\nPay here: {url}\\n\\nAfter payment, send your flyer details, logo, photos, menu, or sample flyer.' Order status=pending_payment.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Start guest order with no checkout URL configured",
        "Stripe/Razorpay URL not configured",
        "1. Trigger start_guest_order.",
        "Reply: 'Payment link is not configured yet. I will send it here when it is ready.' Order still saved as pending_payment.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Start guest order is idempotent for same sender+chat",
        "Existing pending_payment order for sender X1",
        "1. Trigger start_guest_order again.",
        "No new order created. Same pending-payment reply returned.",
        "P0", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Activate order via payment_reference",
        "Pending order order_id=GO-001 for sender X1",
        "1. Run manage-flyer-guest-order --activate --order-id GO-001 --payment-reference pi_123 --sender-phone X1.",
        "Order status -> paid. Reply sent: 'Flyer Studio\\n------------\\nPayment received for GO-001. Send the flyer details now.\\n\\nYou can send text, logo, photos, menu, or a sample flyer.'",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Activate idempotent: same payment_reference replayed",
        "Order GO-001 already paid with payment_reference=pi_123",
        "1. Run activate again with same args.",
        "Existing paid reply returned. No state change.",
        "P0", "Edge", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Activate fails: missing payment_reference",
        "Pending order",
        "1. Run activate without --payment-reference.",
        "Result: detail='payment_reference_required'. Order unchanged.",
        "P1", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Activate fails: payment_reference already used on different order",
        "Order GO-002 holds payment_reference=pi_123",
        "1. Activate GO-001 with --payment-reference pi_123.",
        "detail='payment_reference_already_used'. Order unchanged.",
        "P0", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Activate fails on used order",
        "Order GO-001 status=used",
        "1. Activate with new payment_reference.",
        "detail='cannot_activate_used'. Order unchanged.",
        "P1", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reserve order against a project",
        "Paid order; new project P1 just created",
        "1. Run --reserve --project-id P1 --sender-phone X1 --chat-id C1.",
        "Status -> reserved; reserved_project_id=P1.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Release reserved order on project failure",
        "Reserved order; project P1 fails to render",
        "1. Run --release --project-id P1 --sender-phone X1 --chat-id C1.",
        "Status -> paid; reserved_project_id cleared.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Consume order on delivery; remaining count > 0 keeps status paid",
        "Order with flyer_count_purchased=2; flyer_count_used=0; reserved against P1",
        "1. Run --consume --project-id P1.",
        "Status remains paid (1 of 2 used); used_project_ids contains P1.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Consume order: count exhausted -> status used",
        "Single-flyer order reserved against P1",
        "1. Run --consume --project-id P1.",
        "Status -> used. Future find_paid_order_by_sender returns None.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Consume idempotent for same project_id",
        "Order status=used after first consume",
        "1. Run --consume --project-id P1 again.",
        "No-op; no error. used_project_ids unchanged.",
        "P1", "Edge", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "find-paid: paid order discovered by sender_phone+chat",
        "Paid order for X1 in C1",
        "1. Run --find-paid --sender-phone X1 --chat-id C1.",
        "Result contains order JSON. Used by hooks to set force_new=True path.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    return out


def admin_dashboard_scenarios():
    A, name = "A6", "Admin Dashboard"
    out, i = [], 1
    out.append(row(i, A, name,
        "Operator login + OTP",
        "Operator credentials valid; OTP service available",
        "1. Navigate to /portal/flyer-admin.\n2. Sign in.\n3. Enter OTP.",
        "Session established. require_auth endpoints accessible; require_fresh_otp endpoints require recent OTP step (<5 min).",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "GET /flyer/summary returns expected segments",
        "Operator session",
        "1. Fetch summary via UI or curl.",
        "Returns {segments: {free_trial, paid, payment_pending, inactive, one_time}, total_customers, active_projects, stuck_projects, guest_orders, campaign_asset}.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "List customers filtered by segment",
        "Operator session",
        "1. GET /flyer/customers?segment=free_trial.",
        "Returns only free-trial customers; pagination works; max 300 results.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Search customer by name/phone",
        "Operator session; multiple customers",
        "1. GET /flyer/customers?query=Triveni.",
        "Returns customers whose id, business_name, public_phone, business_whatsapp_number, or authorized_request_numbers contain 'Triveni' case-insensitively.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Customer detail page",
        "Operator session",
        "1. GET /flyer/customers/{id}.",
        "Returns customer record + linked active/recent projects + usage_events; raw profile JSON included.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Extend trial adds bonus flyers",
        "Active trial customer; OTP fresh",
        "1. POST /flyer/customers/{id}/extend-trial body {reason:'support escalation', extra_flyers:2}.",
        "trial_bonus_flyers incremented by 2; timestamped backup created; audit entry flyer.customer.extend_trial with IP+UA.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Extend trial fails without fresh OTP",
        "Operator session > 5 min since OTP",
        "1. POST extend-trial.",
        "HTTP 401 or OTP-required response.",
        "P0", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reset trial releases reserved/used quota events",
        "Trial customer with usage_events: 2 reserved, 1 used",
        "1. POST /flyer/customers/{id}/reset-trial body {reason:'support reset'}.",
        "Three 'released' events appended; trial quota effectively zeroed; backup file created.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Reset trial idempotent on re-run",
        "Already-reset customer",
        "1. POST reset-trial again.",
        "No new released events appended (latest events for each reservation_id are already 'released').",
        "P1", "Edge", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Extend trial rejects extra_flyers out of range",
        "Operator session",
        "1. POST extend-trial extra_flyers=0 or 101.",
        "HTTP 422 validation error.",
        "P1", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Projects list filterable by status",
        "Operator session",
        "1. GET /flyer/projects?status=awaiting_final_approval.",
        "Returns matching projects sorted by updated_at desc; max 300.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Guest orders list",
        "Operator session",
        "1. GET /flyer/guest-orders.",
        "Returns up to 300 guest orders sorted by updated_at desc.",
        "P1", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign preview (text targets, dry_run=true default)",
        "Operator session",
        "1. POST /flyer/campaigns/preview body {targets_text:'+1 555 010 0001\\n+1 555 010 0002', dry_run:true, include_paid:false, reason:'test'}.",
        "Returns {valid_targets, invalid, duplicate_count} without sending; no audit-as-send entry.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign CSV preview missing 'phone' header",
        "Operator session",
        "1. POST /flyer/campaigns/preview-csv with CSV lacking phone column.",
        "HTTP 422 'CSV missing required header phone'.",
        "P0", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign CSV over 512 KB rejected",
        "Operator session",
        "1. Upload a 600 KB CSV.",
        "HTTP 413.",
        "P1", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign CSV with formula injection rejected",
        "Operator session; CSV cell starting with '=' or '@' (non-phone)",
        "1. Upload CSV.",
        "HTTP 422 formula-injection error. Phone cells starting '+' with digits/spaces/dashes are exempted.",
        "P0", "Negative", "Admin UI",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign send actually delivers (dry_run=false)",
        "Fresh OTP; valid targets",
        "1. POST /flyer/campaigns/send body {targets_text:'+1 555 010 0001', dry_run:false, reason:'launch'}.",
        "send-flyer-campaign script invoked per target; counts of sent/failed returned; audit entry flyer.campaign.send recorded with reason + target_count.",
        "P0", "Happy", "Admin UI",
        ""
    )); i += 1
    return out


def cf_router_scenarios():
    A, name = "A7", "cf-router Routing"
    out, i = [], 1
    out.append(row(i, A, name,
        "Owner self-chat with #XXXXX approval code bypasses LLM",
        "Owner chat with active 5-char approval code pending",
        "1. From owner phone, reply '#ABC12 yes'.",
        "F8 path intercepts before any flyer logic; approval marked applied; no flyer routing.",
        "P0", "Edge", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Campaign CTA short-circuits dispatch chain",
        "is_flyer_campaign_cta=True for inbound",
        "1. Customer taps campaign button.",
        "Only _try_flyer_campaign_cta_intercept runs; subsequent intercepts (intake/account/etc.) NOT evaluated.",
        "P0", "Edge", "WhatsApp",
        "Line 149-153 in hooks.py: returns CTA result or None; never falls through to intake."
    )); i += 1
    out.append(row(i, A, name,
        "Intake intercept fires for unregistered sender with vague start",
        "No customer; sends 'Make flyer'",
        "1. Sender X3 (no record) sends 'make flyer'.",
        "is_vague_flyer_start=True; role != owner; new intake session created with source=start_trial. Reply: language menu.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Existing-onboarding intercept handles in-progress session",
        "Sender has in-progress onboarding session",
        "1. Send 'BACK'.",
        "_try_flyer_existing_onboarding_intercept handles BACK; advances state machine accordingly.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Active customer + vague start receives starter brief, not project",
        "Active trial customer",
        "1. Send 'I want a flyer'.",
        "Starter brief intercept fires; audit reason='flyer_starter_brief'; action=skip.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Payment_pending customer + vague start gets account-not-active reply",
        "Customer status=payment_pending",
        "1. Send 'create flyer'.",
        "Reply: 'Your account is waiting for payment confirmation. I saved your account details, but flyer generation starts after activation.' audit reason='flyer_customer_not_active'.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Suspended customer + vague start",
        "Customer status=suspended",
        "1. Send 'create flyer'.",
        "Reply: 'This Flyer Studio account is suspended. Contact Support before creating a new flyer.'",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Cancelled customer + vague start",
        "Customer status=cancelled",
        "1. Send 'create flyer'.",
        "Reply: 'This Flyer Studio account is cancelled. Contact Support or restart setup before creating a new flyer.'",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Account command STATUS",
        "Active customer",
        "1. Send 'status'.",
        "Reply: 'Account: {id}\\nStatus: {status}\\nPlan: {plan_id}\\nUsage this period: {used} used, {remaining} remaining.{pending}{trial_cta}'.",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Account command HELP differs by status",
        "Trial, payment_pending, and active customer separately",
        "1. From each, send 'help'.",
        "Three distinct help replies as documented (trial CTA list, payment-pending wait, active commands).",
        "P1", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Non-admin attempts mutating command",
        "Active customer; sender not in authorized_request_numbers",
        "1. Send 'add authorized 555 010 9999'.",
        "Reply: 'Only the business WhatsApp number or account owner can change account settings.'",
        "P0", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Two-step CONFIRM UPDATE flow for remove_authorized",
        "Authorized admin sends 'remove authorized 555 010 0009'",
        "1. Send command.\n2. Reply 'CONFIRM UPDATE' from same admin number.",
        "Step 1 reply: 'Please reply CONFIRM UPDATE to apply this account change.' Step 2 applies the removal.",
        "P0", "Happy", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "CONFIRM UPDATE from wrong admin",
        "Pending update initiated by admin A1; A2 also authorized",
        "1. From A2 reply 'CONFIRM UPDATE'.",
        "Reply: 'Please confirm from the same admin number that requested this update.' Pending update untouched.",
        "P1", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "CONFIRM UPDATE with no pending update",
        "No pending update on account",
        "1. Send 'CONFIRM UPDATE'.",
        "Reply: 'No pending account update.'",
        "P2", "Negative", "WhatsApp",
        ""
    )); i += 1
    out.append(row(i, A, name,
        "Audit log dispatcher_routed entry written for every classified inbound",
        "Any classified inbound",
        "1. Send any message that triggers an intercept.",
        "decisions.log NDJSON record with type='cf_router_intercepted' (or dispatcher_routed) including {reason, action, chat_id, message_id}. Layer 0 of test pyramid.",
        "P0", "Edge", "WhatsApp",
        ""
    )); i += 1
    return out


def all_scenarios():
    return (
        onboarding_scenarios()
        + text_mode_scenarios()
        + image_reference_scenarios()
        + active_project_scenarios()
        + guest_order_scenarios()
        + admin_dashboard_scenarios()
        + cf_router_scenarios()
    )


def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)


def apply_body_style(cell):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.font = Font(name="Calibri", size=10)


def write_sheet(ws, rows):
    ws.append(HEADERS)
    for c in ws[1]:
        apply_header_style(c)
    for r in rows:
        ws.append(r)
    widths = [16, 18, 38, 36, 50, 60, 10, 11, 12, 32]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    last_col = len(HEADERS)
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=last_col):
        for c in row_cells:
            if c.row > 1:
                apply_body_style(c)
            c.border = border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def build_workbook(out_path: Path):
    wb = Workbook()
    master = wb.active
    master.title = "All Scenarios"
    rows = all_scenarios()
    write_sheet(master, rows)
    by_area = {code: [] for code, _ in AREAS}
    for r in rows:
        code = r[0].split("-")[1]
        by_area[code].append(r)
    for code, area_name in AREAS:
        safe_name = area_name.replace("/", "-")
        sheet_title = f"{code} {safe_name}"[:31]
        ws = wb.create_sheet(title=sheet_title)
        write_sheet(ws, by_area[code])
    summary = wb.create_sheet(title="README", index=0)
    summary["A1"] = "Flyer Studio - Manual QA Test Scenarios"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Source"
    summary["B3"] = "Branch codex/flyer-edit-pipeline (PRs #100, #101, #102)"
    summary["A4"] = "Generated"
    summary["B4"] = "2026-05-19"
    summary["A6"] = "Sheets"
    summary["A6"].font = Font(bold=True)
    summary["A7"] = "All Scenarios"; summary["B7"] = f"{len(rows)} total"
    for idx, (code, area_name) in enumerate(AREAS, start=8):
        safe_name = area_name.replace("/", "-")
        summary[f"A{idx}"] = f"{code} {safe_name}"
        summary[f"B{idx}"] = f"{len(by_area[code])} scenarios"
    summary["A16"] = "Columns"
    summary["A16"].font = Font(bold=True)
    legend = [
        ("ID", "FS-<Area>-<NNN> stable identifier"),
        ("Area", "One of 7 functional areas"),
        ("Scenario", "Short test title"),
        ("Preconditions", "Required state before steps"),
        ("Steps", "Numbered tester actions"),
        ("Expected Result", "Observable outcome including reply text where exact"),
        ("Priority", "P0 = blocker, P1 = important, P2 = nice-to-have"),
        ("Type", "Happy / Negative / Edge"),
        ("Channel", "WhatsApp inbound or Admin UI"),
        ("Notes", "Cross-references, file paths, regression links"),
    ]
    for idx, (k, v) in enumerate(legend, start=17):
        summary[f"A{idx}"] = k
        summary[f"B{idx}"] = v
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 80
    wb.save(out_path)
    return len(rows)


if __name__ == "__main__":
    out_path = Path(__file__).parent / "flyer-studio-qa-scenarios.xlsx"
    count = build_workbook(out_path)
    print(f"Wrote {count} scenarios to {out_path}")
